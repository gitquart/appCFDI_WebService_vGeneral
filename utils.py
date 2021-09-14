from cfdiclient import Autenticacion,Fiel,SolicitaDescarga,VerificaSolicitudDescarga,DescargaMasiva,Validacion
import os
import base64
import zipfile
from xml.etree import ElementTree as ET
import openpyxl as excelpy
from InternalControl import cInternalControl
import postgresql as bd
import time
import tkinter.messagebox as tkMessageBox
import smtplib,ssl
from email.message import EmailMessage
import datetime
import threading

#Important information for this code
#--------------------------------------------------------------------------
#Folder and file names: [RFC]_[TIPO]_[FECHAS]_[ID]
#lsFolder elements (by order):[rfc_solicitante,tipo,fechaCompleta,fileName] 

objControl=cInternalControl()
formatDateTime='%Y-%m-%d %H:%M'
#They are filled wit datos.text
rfcFromFile=''
FIEL_PAS = ''
rfc_emisor= ''
rfc_receptor=''
rfc_solicitante=''
#End-They are filled wit datos.text
FIEL_KEY = ''
FIEL_CER = ''
fiel=''
VERSION=''
ID_CURRENT_SOLICITUD=''


def sendMail(mailToAuthorize):
    """
    To get this data: Go to Quart mail on web, go to Settings, View all outlook settings,Sync email, find smtp....
    Server name: smtp.office365.com
    Port: 587
    Encryption method: STARTTLS
    """
    mfrom='ulyses@quart.com.mx'
    mto='ulyses@quart.com.mx'
    msg = EmailMessage()
    msg.set_content(f'Please authorize {mailToAuthorize} in database.')
    msg['Subject'] = f'Please authorize CFDI user'
    msg['From'] = mfrom
    msg['To'] = mto
    context=ssl.create_default_context()
    try:
        server = smtplib.SMTP(host='smtp.office365.com',port=587)
        server.starttls(context=context)
        server.login(mfrom,'Feaconfdefoca33')
        server.send_message(msg)
        server.quit()
    except NameError as error:
        showMessage(f'Su usuario ha sido registrado exitosamente, pero ha ocurrido un error al mandar el correo: {error}. ')   
 
def showMessage(title,content):
    tkMessageBox.showinfo(title=title,message=content) 

def validateFIELFiles(directory):
    global rfcFromFile,rfc_emisor,rfc_receptor,rfc_solicitante,fiel,FIEL_CER,FIEL_KEY,FIEL_PAS
    numFiles=len(os.listdir(directory))
    for file in os.listdir(directory):
        if file == "datos.txt":
            datostxt = open(directory+'/datos.txt', 'r')
            Lines = datostxt.readlines()
            count=1
            for line in Lines:
                if line!="\n":
                    if count==1:
                        rfcFromFile=line.strip()
                        rfc_emisor= rfcFromFile
                        rfc_receptor=rfcFromFile
                        rfc_solicitante=rfcFromFile
                        count+=1
                        continue
                    if count==2:  
                        #Don't strip the line for password, we had a case of "CURP ", which space is needed  
                        FIEL_PAS=line
                        break

        if file.endswith('.key'):
            FIEL_KEY=file
            continue
        if file.endswith('.cer'):
            FIEL_CER=file
            continue 
    if FIEL_CER!='' and FIEL_KEY!='':    
        cer_der = open(directory+'/'+FIEL_CER, 'rb').read()
        key_der = open(directory+'/'+FIEL_KEY, 'rb').read()  
        fiel = Fiel(cer_der, key_der, FIEL_PAS)
    
    return numFiles
  
def autenticacion():
    auth = Autenticacion(fiel)
    token = auth.obtener_token()

    return token

def solicitaDescarga(fecha_inicial,fecha_final,directory,tipo,fechaCompleta,Version,window):
    global VERSION,ID_CURRENT_SOLICITUD
    #Get the version so it can be stored in utils.py for the rest of methods
    VERSION=Version
    #Ejemplo de respuesta  {'mensaje': 'Solicitud Aceptada', 'cod_estatus': '5000', 'id_solicitud': 'be2a3e76-684f-416a-afdf-0f9378c346be'}
    try:
        res=validateFIELFiles(directory)
    except:
        result=[0,'Hubo un error con los archivos, favor de verificar que los archivos CER,KEY o datos.txt sean los correctos']    
        return result
    if res>0:
        #Officialy it is processing
        strTitle=window.title()
        strTitle+='- CFDI request'
        window.title(strTitle)
        if VERSION=='SQL':
            #1)Check if the user is in table usuario
            resultSet=bd.getQueryOrExecuteTransaction("select * from usuario where rfc_solicitante='"+rfc_solicitante+"' ;")
            if len(resultSet)==0:
                #insert the rfc_solicitante in usuario
                resultSet=bd.getQueryOrExecuteTransaction("insert  into usuario (rfc_solicitante) values ('"+rfc_solicitante+"') RETURNING id;")
            #2)Check if the request exists already, otherwise insert it
            id_usuario=''
            for item in resultSet[0]:
                id_usuario=str(item)
                break 
            #postresql date= 'yyyy-mm-dd' 
            fechaInicial=str(fecha_inicial.year)+'-'+str(fecha_inicial.month)+'-'+str(fecha_inicial.day)  
            fechaFinal= str(fecha_final.year)+'-'+str(fecha_final.month)+'-'+str(fecha_final.day)
            query="select * from solicitud where id_usuario="+id_usuario+" and fechainicio='"+fechaInicial+"' and fechafin='"+fechaFinal+"' and tipo='"+tipo+"' ; "
            resultSet=bd.getQueryOrExecuteTransaction(query)
            if len(resultSet)==0:
                cmd="insert into solicitud (fechainicio,fechafin,conteo,tipo,id_usuario) values ('"+fechaInicial+"','"+fechaFinal+"',0,'"+tipo+"',"+id_usuario+") returning id;"
                resultSet=bd.getQueryOrExecuteTransaction(cmd)
                for item in resultSet[0]:
                    ID_CURRENT_SOLICITUD=str(item)
                    break 
            else:
                #If the "solicitud" already exists, then return a message "La solicitud sólo puede ser ejecutada 1 vez"
                return [1,'Mensaje de base de datos: La petición de este CFDI ya ha sido solicitada 1 vez']    
        lsfolderName=[]
        lsfolderName.append(rfc_solicitante)
        lsfolderName.append(tipo)
        lsfolderName.append(fechaCompleta)
        descarga = SolicitaDescarga(fiel)
        token = autenticacion()
        result=''
        print('Start - CFDI Webservice request')
        if lsfolderName[1]=='Emisor':
            # Emitidos
            result = descarga.solicitar_descarga(token, rfc_solicitante, fecha_inicial, fecha_final, rfc_emisor=rfc_emisor)
        else:    
            # Recibidos
            result = descarga.solicitar_descarga(token, rfc_solicitante, fecha_inicial, fecha_final, rfc_receptor=rfc_receptor)
        
        #Here the ID of the request is done, so let's wait TIME_REQUEST_MINS minutes lo let SAT set state 3,
        #after TIME_REQUEST_MINS minutes let's chake the state and it should be 3 and correct
        time.sleep(objControl.TIME_FOR_REQUEST) #600 secs = 10 mins, 2400 secs= 40 mins
        res=verificaSolicitudDescarga(result['id_solicitud'],directory,lsfolderName,window)
        return res
        
    else:
        return [0,"El directorio no contiene archivos FIEL"]   

def verificaSolicitudDescarga(id_solicitud,directory,lsFolderName,window):
    if id_solicitud!='':
        #Ejemplo re respuesta  {'estado_solicitud': '3', 'numero_cfdis': '8', 'cod_estatus': '5000', 'paquetes': ['a4897f62-a279-4f52-bc35-03bde4081627_01'], 'codigo_estado_solicitud': '5000', 'mensaje': 'Solicitud Aceptada'}   
        v_descarga = VerificaSolicitudDescarga(fiel)
        token = autenticacion()
        result = v_descarga.verificar_descarga(token, rfc_solicitante, id_solicitud)
        #I add some seconds before it gets the result for "verificar"
        print('End of CFDI Web service request')
        print('Number of CFDI:',result['numero_cfdis'],' from ID request :',str(id_solicitud))
        #Change Title again
        strTitle=window.title() 
        strTitle=str(strTitle).replace('- CFDI request','')  
        window.title(strTitle)   
        if (int(result['numero_cfdis'])>0):
            lsFolderName.append(result['paquetes'][0])
            res=descargarPaquete(result['paquetes'],directory,lsFolderName)
            if int(res[0])==1:
                if VERSION=='SQL':
                    return [1,'Procesamiento exitoso, el archivo ZIP con CFDI se descargó en '+directory+'/'+result['paquetes'][0]+' y se cargaron los registros en la base de datos.']
                else:
                    return [1,'Procesamiento exitoso, el resultado se descargó en '+directory+'/'+result['paquetes'][0]+' (zip y xlsx) ']
                
            else:
                return res    
        else:
            mensaje=''
            if VERSION=='SQL':
                cmd="update solicitud set conteo=1 where id="+ID_CURRENT_SOLICITUD+";"
                bd.getQueryOrExecuteTransaction_NoReturning(cmd)
                mensaje='El paquete no trae CFDI y se ha registrado la operación en base de datos, respuesta de web service:'
            else:
                mensaje='El paquete no trae CFDI, respuesta de web service:\n'

            return [
                     0,str(mensaje)+'\n'+
                     'Estado de solicitud:'+result['estado_solicitud']+'\n'+
                     'Código de estado:'+result['cod_estatus']+'\n'+
                     'Número de CFDI:'+result['numero_cfdis']+'\n'+
                     'Código de estado de la solicitud:'+str(result['codigo_estado_solicitud'])+'\n'+
                     'Mensaje:'+result['mensaje']+'\n\n'+
                     '------Estados de solicitud--------\n'+
                     '1 - Aceptada\n'+
                     '2 - En proceso\n'+
                     '3 - Terminada\n'+
                     '4 - Error\n'+
                     '5 - Rechazada\n'+
                     '6 - Vencida\n\n'+
                     'Solicitud original:'+str(id_solicitud)

                     ]  
    else:
        return [0,'No se encontró Solicitud'] 
      
def validaEstadoDocumento(rfc_emisor,rfc_receptor,uuid,total):
    """
    validaEstadoDocumento
    Parms: All input parameters are text 
    """
    validacion = Validacion()
    estado=''
    try:
        estado = validacion.obtener_estado(rfc_emisor, rfc_receptor, total, uuid)
    except:
        print('First Attempt to Reconnect to Valida estado in 10 seconds')
        time.sleep(10)
        try:
            estado = estado = validacion.obtener_estado(rfc_emisor, rfc_receptor, total, uuid)
        except:    
            print('Second Attempt to Reconnect to Valida estado in 10 seconds')
            time.sleep(10)
            try:
                estado = estado = validacion.obtener_estado(rfc_emisor, rfc_receptor, total, uuid)
            except:
                print('First and Second to reconnect attempt failed.')    
            

    return estado

def descargarPaquete(id_paquete,directory,lsFolderName):
    #ejemplo de respuesta # {'cod_estatus': '', 'mensaje': '', 'paquete_b64': 'eyJhbG=='} 
    descarga = DescargaMasiva(fiel)
    token = autenticacion()
    result = descarga.descargar_paquete(token, rfc_solicitante, id_paquete[0])
    paquete=result['paquete_b64']
    if paquete is not None:
        #if paquete is not None, the create the folder where zip and xls will be saved
        folderAndFileName='_'.join(lsFolderName)
        ZipExcelDir=directory+'/'+folderAndFileName
        if not os.path.isdir(ZipExcelDir):
            os.mkdir(ZipExcelDir)
        readBase64FromZIP(paquete,folderAndFileName,ZipExcelDir)
        print('Downloading package with CFDI and processing...')
        if VERSION=='EXCEL':
            extractAndReadZIP(ZipExcelDir,folderAndFileName+'.zip',rfc_solicitante)      
        else:
            extractAndReadZIP_SQL(ZipExcelDir,folderAndFileName+'.zip',rfc_solicitante)

        return [1]
    else:
        return [0,'No se descargó CFDI: '+result['mensaje']]

def readBase64FromZIP(file,folderAndFileName,directory): 
    """
    readBase64FromZIP: Reads the package in base64 from SAT and returns the zip file (the zip file is actually)
    created on the go.
    """
    if file is not None:
        with open(directory+'/'+folderAndFileName+'.zip', 'wb') as result:
            result.write(base64.b64decode(file))
        zip_ref = zipfile.ZipFile(directory+'/'+folderAndFileName+'.zip', 'r')
        zip_ref.close()

def extractAndReadZIP_SQL(directory,zipToRead,rfc_solicitante):
    objControl=cInternalControl()
    #Change / to \\ if neccesary
    separationFolder=''
    if objControl.testingMode:
        separationFolder='\\'
    else:
        separationFolder='/'    
    myZip=zipfile.ZipFile(directory+separationFolder+zipToRead,'r')
    contDocs=0
    #dicTableFields is a dictionary with the following structura key:table, value: list of fields
    dicTableFields={}
    #Dictionaries for every kind of "tipo de comprobante"
    #First, get all the columns for all the tables
    for xml in myZip.namelist():
        doc_xml=myZip.open(xml)
        root = ET.parse(doc_xml).getroot()
        for node in root.iter():
            #Column = attribute , Node= Table
            #Get attributes (columns) of current Node (table) 
            #Split the node (table) name because all come as "{something}Node" and {content} is not needed
            #If the number of nodes > 1 then not get its fields, we only want 1 row
            chunk=str(node.tag).split('}')
            tableName=chunk[1] 
            numOfNodes=len(node.getchildren())
            if (numOfNodes<2) or (numOfNodes>1 and tableName=='Comprobante'):
                chunk=str(node.tag).split('}')
                tableName=chunk[1]  
                #As all the fields will be in one single sheet, it can occur two fields with the same
                #name ex: Rfc and Rfc (Emisor and recipient), then it's needed to add prefix tableName
                for attr in node.attrib:
                    fieldName=tableName+'_'+attr
                    if tableName not in dicTableFields:
                        dicTableFields[tableName]=[fieldName]
                    else:
                        if fieldName not in dicTableFields[tableName]:
                            dicTableFields[tableName].append(fieldName)            
            #End of node iteration 

    #Second, when got all fields from all xml, print them in spread sheet
    lsFields=[] 
    #Add extra fields here
    lsFields.append('nombreArchivo')
    lsFields.append('mes')
    lsSource=[]
    if len(objControl.lsCustomFields)==0:
        lsSource=dicTableFields   
    else:
        lsSource=objControl.lsCustomFields    

    #I append insetad of just assign the list, because I need the column "mes" in the very beginning
    for field in lsSource:
        lsFields.append(field)

    for field in objControl.lsRemove:
        if field in lsFields:
            lsFields.remove(field)     
    
    #Add id_solicitud at the end
    lsFields.append('id_solicitud')
    #lsFieldsSQL is the list which contains the name of fields as shown in database, all the fields in tables are
    # the same as fields but in lower case, this means: "Comprobante_Fecha" turns to "comprobante_fecha"      
    lsFieldsSQL = [x.lower() for x in lsFields] 
    #Convert lsFieldsSQL into the way they will appear in statement : insert into... -> (field1,field2,...) 
    fieldsInStatement='('+','.join(lsFieldsSQL)+')'
    #Third, read information and insert where belongs 
    #Conclusiones: 
    # getroot() : Gets the root of the xml, then use getRoot to get "Comprobante"
    # root.find(.//...): gets any node inside the root, use this to any other node except the root
    for xml in myZip.namelist():
        #Get field TipoDeComprobante to knwo where sheet to print
        doc_xml=myZip.open(xml)
        root = ET.parse(doc_xml).getroot()
        lsRfcTable=['Emisor','Receptor']
        for item in lsRfcTable:
            node=returnFoundNode(root,item)
            if len(node)>0:
                rfc_value=node[0].get('Rfc')
                if rfc_value==rfc_solicitante:
                    if ((root.get('TipoDeComprobante')=='I' or root.get('TipoDeComprobante')=='Ingreso' or root.get('TipoDeComprobante')=='ingreso' ) or 
                        (root.get('TipoDeComprobante')=='E') or (root.get('TipoDeComprobante')=='Egreso') or (root.get('TipoDeComprobante')=='egreso') ):
                        tableSQL=item
                    elif  (root.get('TipoDeComprobante')=='P' or root.get('TipoDeComprobante')=='Pago' or root.get('TipoDeComprobante')=='pago'):
                        tableSQL='Pago' 
                    else:
                        tableSQL='Pago'
                    break                

        #Start to read the fields from lsFields=[]
        #Example of a field in lsFields : "Comprobante_Version" -> "tableName_Field"
        #One row per xml
        lsRow=[]
        #The field leads all the insertion
        #Algorithm of reading fields
        for field in lsFields:
            #Cases
            if field=='nombreArchivo':
                lsRow.append(xml)
                continue
            if field=='mes':
                fechaFactura=root.get('Fecha')
                monthWord=returnMonthWord(int(fechaFactura.split('-')[1]))
                lsRow.append(monthWord)
                continue
            if field=='id_solicitud':
                # Add id_solicitud value
                #For test case only (when running from main.py)
                #ID_CURRENT_SOLICITUD='6'
                #End "For test case only..."
                lsRow.append(ID_CURRENT_SOLICITUD) 
                continue
            #Rest of cases
            chunks=field.split('_')
            table=chunks[0]
            column=chunks[1]
            if table=='Comprobante':  
                addColumnIfFound_SQL(root,column,lsRow,'0')  
            else:
                #Find the right prefix for table
                lsNode=returnFoundNode(root,table)
                if len(lsNode)==1:
                    addColumnIfFound_SQL(lsNode[0],column,lsRow,'0')
                elif len(lsNode)>1:
                    #More than 1 table_column (Node) found with the same name in XML
                    bTableWithField=False
                    for node in root.findall('.//'+objControl.prefixCFDI+table):
                        if len(node.attrib)>0: 
                            #If this table has attributes, read it, other wise skip it becase
                            #if the column doesn't have fields, it means it holds children
                            addColumnIfFound_SQL(node,column,lsRow,'0')
                            bTableWithField=True
                    if not bTableWithField:
                        #The table exists, but it doesn't have the current field
                        lsRow.append('0') 

                else:
                    #No table name found
                    lsRow.append('0')  
            #End of field iteration

        #Append the whole xml in a single row in sql 
        #Convert lsRow into a correct value list for SQL
        lsFieldsNotToMatch=['comprobante_subtotal','impuestos_totalimpuestosretenidos','impuestos_totalimpuestostrasladados',
                        'comprobante_total','comprobante_descuento','id_solicitud']
        lsFieldsToMatch=[]
        for field in lsFieldsSQL:
            if field not in lsFieldsNotToMatch:
                lsFieldsToMatch.append(field)

        for field in lsFieldsToMatch:                
            transforValuesToSQLFormat(field,lsFieldsSQL,lsRow)
        valuesInSatement=",".join(lsRow)          
        finalCmd="insert into "+tableSQL+" "+fieldsInStatement+" values ("+valuesInSatement+") ;"  
        bd.getQueryOrExecuteTransaction_NoReturning(finalCmd)        
        contDocs+=1
        #End of each document (xml) iteration in a zip
        

    #All xml processed at this point    
    cmd="update solicitud set conteo=1 where id="+ID_CURRENT_SOLICITUD+";"
    bd.getQueryOrExecuteTransaction_NoReturning(cmd)
    print('Files processed in ZIP file:',str(contDocs)) 

def extractAndReadZIP(directory,zipToRead,rfc_solicitante,testingMode=False):
    separationFolder=''
    if testingMode:
        separationFolder='\\'
    else:
        separationFolder='/'    
    myZip=zipfile.ZipFile(directory+separationFolder+zipToRead,'r')
    #The zip's file name will be the name of excel file name, like the "Database"
    excel_fileName=os.path.splitext(os.path.split(myZip.filename)[1])[0]+'.xlsx'
    #Creating the workbook (database)
    #Create the sheets: Ingreso_Egreso,Pago,Resto
    wb=excelpy.Workbook() 
    #let's create the sheets in this order: Emisor, Receptor, Pago_Emisor,Pago_Receptor
    #Sheet1 is the first, so rename it.
    pago_sheet = wb['Sheet']
    pago_sheet.title = 'Emisor'
    wb.create_sheet('Receptor')
    wb.create_sheet('Pago_Emisor')
    wb.create_sheet('Pago_Receptor')
    
    wb.save(directory+'/'+excel_fileName)
    contDocs=0
    #dicTableFields is a dictionary with the following structura key:table, value: list of fields
    dicTableFields={}
    #Dictionaries for every kind of "tipo de comprobante"
    #First, get all the columns for all the tables
    for xml in myZip.namelist():
        doc_xml=myZip.open(xml)
        root = ET.parse(doc_xml).getroot()
        for node in root.iter():
            #Column = attribute , Node= Table
            #Get attributes (columns) of current Node (table) 
            #Split the node (table) name because all come as "{something}Node" and {content} is not needed
            #If the number of nodes > 1 then not get its fields, we only want 1 row
            chunk=str(node.tag).split('}')
            tableName=chunk[1] 
            numOfNodes=len(list(node))
            #some "numOfNodes" are Zero, therefore I compare numOfNodes < 2 so I can get 0 and 1
            if (numOfNodes<2) or (numOfNodes>1 and tableName=='Comprobante'):
                chunk=str(node.tag).split('}')
                tableName=chunk[1]  
                #As all the fields will be in one single sheet, it can occur two fields with the same
                #name ex: Rfc and Rfc (Emisor and recipient), then it's needed to add prefix tableName
                for attr in node.attrib:
                    fieldName=tableName+'_'+attr
                    if tableName not in dicTableFields:
                        dicTableFields[tableName]=[fieldName]
                    else:
                        if fieldName not in dicTableFields[tableName]:
                            dicTableFields[tableName].append(fieldName)            
            #End of node iteration 

    #Second, when got all fields from all xml, print them in spread sheet
    lsFields=[] 
    #Start -  Add extra fields in the beginning
    lsFields.append('nombreArchivo')
    lsFields.append('mes')
    #End    - Add extra fields in the beginning

    lsSource=[]
    if len(objControl.lsCustomFields)==0:
        lsSource=dicTableFields   
    else:
        lsSource=objControl.lsCustomFields    

    #I append instead of just assign the list, because I need the column "mes" in the very beginning
    for field in lsSource:
        lsFields.append(field)

    for field in objControl.lsRemove:
        if field in lsFields:
            lsFields.remove(field)     

    #Start - Add extra fields at the end
    lsFields.append('Estatus')
    lsFields.append('Fecha/Hora de Consulta')
    #End - Add extra fields at the end 
    
    #Print all lsFields on excel workbook
    for sheet in wb.sheetnames:
        wb[sheet].append(lsFields)     
    wb.save(directory+'/'+excel_fileName)     
               
  
    #Third, read information and insert where belongs 
    #Conclusiones: 
    # getroot() : Gets the root of the xml, then use getRoot to get "Comprobante"
    # root.find(.//...): gets any node inside the root, use this to any other node except the root
    for xml in myZip.namelist():
        #I need the status for every xml (each cfdi) hence a declare the following fields:
        #Emisor_Rfc,Receptor_Rfc,TimbreFiscalDigital_UUID,Comprobante_Total
        vEmisorRfc=None
        vReceptorRfc=None
        vTimbreFiscal=None
        vComprobanteTotal=None
        #Get field TipoDeComprobante to knwo where sheet to print
        #"Resto" is the default spread sheet
        doc_xml=myZip.open(xml)
        root = ET.parse(doc_xml).getroot()
        lsRfcTable=['Emisor','Receptor']
        sheetPrint='Nada'
        for item in lsRfcTable:
            node=returnFoundNode(root,item)
            if len(node)>0:
                rfc_value=addColumnIfFound(node[0],None,None,'look',['Rfc','rfc'])
                if rfc_value==rfc_solicitante:
                    tipoComprobante=addColumnIfFound(root,None,None,'look',['TipoDeComprobante','tipoDeComprobante'])
                    for possibleValue in ['Ingreso','ingreso','I','i','E','Egreso','egreso','e']:
                        if tipoComprobante==possibleValue:
                            sheetPrint=item
                            break
                    for possibleValue in ['P','p','Pago','pago']:
                        if tipoComprobante==possibleValue:
                            #If "Pago" then it's clear it goes to Pago sheet, so here now divide Pago into
                            #"Pago_Emisor", "Pago_Receptor"
                            if item=='Emisor':
                                sheetPrint='Pago_Emisor'
                            else:
                                sheetPrint='Pago_Receptor'    
                            break
                    if sheetPrint=='Nada': 
                        #If "Nada" it means it goes to Pago, so here now divide Pago into
                        #"Pago_Emisor", "Pago_Receptor"
                        if item=='Emisor':
                            sheetPrint='Pago_Emisor'
                        else:
                            sheetPrint='Pago_Receptor'       

        #Start to read the fields from lsFields=[]
        #Example of a field in lsFields : "Comprobante_Version" -> "tableName_Field"
        #One row per xml
        lsRow=[]
        #The field leads all the insertion
        #Algorith of reading fields
        for field in lsFields:
            #Cases
            if field=='nombreArchivo':
                lsRow.append(xml)
                continue
            if field=='mes': 
                fechaFactura=addColumnIfFound(root,None,None,'look',['Fecha','fecha'])
                monthWord=returnMonthWord(int(fechaFactura.split('-')[1]))
                lsRow.append(monthWord)
                continue
            if field == 'Estatus':
                #Look for this fields and save them to validate status:
                #Emisor_Rfc,Receptor_Rfc,TimbreFiscalDigital_UUID,Comprobante_Total
                print(f'Obteniendo estado para {vTimbreFiscal}')
                strStatus=None
                strStatus=validaEstadoDocumento(vEmisorRfc,vReceptorRfc,vTimbreFiscal,vComprobanteTotal)
                data=None
                if strStatus:
                    data=strStatus['estado']
                else:
                    data='No hubo respuesta'
                lsRow.append(data)    
                continue
            if field == 'Fecha/Hora de Consulta': 
                lsRow.append(datetime.datetime.now().strftime(formatDateTime))   
                continue       
            #Rest of cases
            chunks=field.split('_')
            table=chunks[0]
            column=chunks[1]
            if table=='Comprobante':  
                addColumnIfFound(root,column,lsRow,'add',None)
            else:
                #Find the right prefix for table
                lsNode=returnFoundNode(root,table)
                if len(lsNode)==1:
                    addColumnIfFound(lsNode[0],column,lsRow,'add',None)
                elif len(lsNode)>1:
                    #More than 1 table_column found with the same name in XML
                    bTableWithField=False
                    for node in root.findall('.//'+objControl.prefixCFDI+table):
                        if len(node.attrib)>0: 
                            #If this table has attributes, read it, other wise skip it because
                            #if the column doesn't have fields, it means it holds children
                            addColumnIfFound(node,column,lsRow,'add',None)
                            bTableWithField=True
                    if not bTableWithField:
                        lsRow.append(0)    
                        
                else:
                    #No table name found
                    lsRow.append(0)

            #Get values to validate CFDI
            #Look for this fields and save them to validate status:
            #Emisor_Rfc,Receptor_Rfc,TimbreFiscalDigital_UUID,Comprobante_Total
            currentSizeList=None
            currentSizeList=len(lsRow)
            if field=='Emisor_Rfc':
                vEmisorRfc=str(lsRow[currentSizeList-1]) 
                continue  
            if field == 'Receptor_Rfc':
                vReceptorRfc=str(lsRow[currentSizeList-1]) 
                continue
            if field=='TimbreFiscalDigital_UUID':
                vTimbreFiscal=str(lsRow[currentSizeList-1]) 
                continue  
            if field == 'Comprobante_Total':
                vComprobanteTotal=str(lsRow[currentSizeList-1])
                continue           
            #End of field iteration

        #Append the whole xml in a single row            
        wb[sheetPrint].append(lsRow)              
        contDocs+=1
        print('File done:', xml,'...',str(contDocs)) 
        #End of each document (xml) iteration in a zip
        wb.save(directory+'/'+excel_fileName)

    #All xml processed at this point    
    print('Files processed in ZIP file:',str(contDocs)) 

def getAndTransformValue(table,lsRow,lsPossibleValues,typeOfColumn,notFoundValue):
    bFieldFound=False
    for possibleColumn in lsPossibleValues:
        if possibleColumn in table.attrib:
            bFieldFound=True
            if (table.get(possibleColumn)!=""):
                if typeOfColumn=='float':
                    lsRow.append(float(table.get(possibleColumn)))
                else:
                    lsRow.append(table.get(possibleColumn))
            else:
                #Table found, but no column found
                lsRow.append(notFoundValue)
    if not bFieldFound:
        #Table found, but no column found
        lsRow.append(notFoundValue)

def addColumnIfFound(table,column,lsRow,op,lsPossibleColumns):
    """
    addColumnIfFound
    Returns a list with the value added in it if found
    """
    #Add all cases that a column can be called
    notValueFloat=0
    notValueString='No value'
    value=''
    if op=='look':
        for column in lsPossibleColumns:
            if column in table.attrib:
                if table.get(column)!="":
                    value=table.get(column)
                else:
                    value='No value'
                return value    
    else:
        #Float cases
        if column=='Total':
            getAndTransformValue(table,lsRow,[column,'total'],'float',notValueFloat)
        elif column=='TotalImpuestosTrasladados':
            getAndTransformValue(table,lsRow,[column,'totalImpuestosTrasladados'],'float',notValueFloat) 
        elif column=='TotalImpuestosRetenidos':
            getAndTransformValue(table,lsRow,[column,'totalImpuestosRetenidos'],'float',notValueFloat) 
        elif column=='SubTotal':
            getAndTransformValue(table,lsRow,[column,'subTotal'],'float',notValueFloat)  
        elif column=='Importe':
            getAndTransformValue(table,lsRow,[column,'importe'],'float',notValueFloat) 
        elif column=='Impuesto':
            getAndTransformValue(table,lsRow,[column,'impuesto'],'string',notValueString) 
        elif column=='Tasa':
            getAndTransformValue(table,lsRow,[column,'tasa'],'float',notValueFloat)          
        #End of Float cases
        elif column=='mes' or column=='Fecha':
            getAndTransformValue(table,lsRow,[column,'fecha'],'string',notValueString)       
        elif column=='Serie':
            getAndTransformValue(table,lsRow,[column,'serie'],'string',notValueString) 
        elif column=='Folio':
            getAndTransformValue(table,lsRow,[column,'folio'],'string',notValueString) 
        elif column=='MetodoPago':
            getAndTransformValue(table,lsRow,[column,'metodoDePago'],'string',notValueString) 
        elif column=='TipoDeComprobante':
            getAndTransformValue(table,lsRow,[column,'tipoDeComprobante'],'string',notValueString)           
        elif column=='UsoCFDI':
            getAndTransformValue(table,lsRow,[column],'string',notValueString)
        elif column=='LugarExpedicion':
            getAndTransformValue(table,lsRow,[column],'string',notValueString) 
        elif column=='Moneda':
            getAndTransformValue(table,lsRow,[column],'string',notValueString)
        elif column=='TipoCambio':
            getAndTransformValue(table,lsRow,[column],'string',notValueString)
        elif column=='Descuento':
            getAndTransformValue(table,lsRow,[column],'string',notValueString)
        elif column=='FormaPago':
            getAndTransformValue(table,lsRow,[column],'string',notValueString)
        elif column=='CondicionesDePago':
            getAndTransformValue(table,lsRow,[column,'condicionesDePago'],'string',notValueString)
        elif column=='Version':
            getAndTransformValue(table,lsRow,[column,'version'],'string',notValueString)
        elif column=='Rfc':
            getAndTransformValue(table,lsRow,[column,'rfc'],'string',notValueString)
        elif column=='Nombre':
            getAndTransformValue(table,lsRow,[column,'nombre'],'string',notValueString)
        elif column=='UUID':
            getAndTransformValue(table,lsRow,[column],'string',notValueString)                                               
        else:
            getAndTransformValue(table,lsRow,[column],'string',notValueString)
                    
def addColumnIfFound_SQL(table,column,lsRow,notFoundValue):
    if column in table.attrib:
        if (table.get(column)!=""):
            lsRow.append(table.get(column))
        else:    
            lsRow.append('0')
    
    else:
        #Table found, but no column found
        lsRow.append(notFoundValue)            

def returnFoundNode(root,table):
    #returnFoundNode: regresa nodo (tabla) si existe en el XML
    lsNode=[]
    result=[]
    for prefix in objControl.lsPrefix:
        lsNode=root.findall('.//'+prefix+table) 
        if len(lsNode)>0:
            result=lsNode
            break
    #If the code reaches this point, it means the Node doesn't exit in the XML, therefore return an empty list
    return result  

def transforValuesToSQLFormat(field,lsFields,lsValuesToTransform):
    #This method works to match the desired fields
    #Case to match => item is field
    for index,item in enumerate(lsFields):
        if item is field:
            newValue="'"+lsValuesToTransform[index]+"'"
            lsValuesToTransform[index]=newValue
            break

def returnMonthWord(monthNumber):
    monthWord=''
    if monthNumber==1:
        monthWord='Enero'
    if monthNumber==2:
        monthWord='Febrero'
    if monthNumber==3:
        monthWord='Marzo'
    if monthNumber==4:
        monthWord='Abril'
    if monthNumber==5:
        monthWord='Mayo'
    if monthNumber==6:
        monthWord='Junio'         
    if monthNumber==7:
        monthWord='Julio'
    if monthNumber==8:
        monthWord='Agosto'
    if monthNumber==9:
        monthWord='Septiembre'  
    if monthNumber==10:
        monthWord='Octubre'
    if monthNumber==11:
        monthWord='Noviembre'
    if monthNumber==12:
        monthWord='Diciembre'                                     

    return monthWord       

#MÉTODOS DE CONSOLA

def verificaSolicitudDescarga_Consola(id_solicitud,rfcsolicitante,directory):
    #Ejemplo re respuesta  {'estado_solicitud': '3', 'numero_cfdis': '8', 'cod_estatus': '5000', 'paquetes': ['a4897f62-a279-4f52-bc35-03bde4081627_01'], 'codigo_estado_solicitud': '5000', 'mensaje': 'Solicitud Aceptada'}   
    try:
        res=validateFIELFiles(directory)
    except:
        result=[0,'Hubo un error con los archivos, favor de verificar que los archivos CER,KEY o datos.txt sean los correctos']    

    if res>0:    
        v_descarga = VerificaSolicitudDescarga(fiel)
        token = autenticacion()
        #La respuesta tiene "paquetes", esos son ID's
        result = v_descarga.verificar_descarga(token, rfc_solicitante, id_solicitud)

        return result['paquetes']
        
#Batch processes (Multi threading)

def extractAndReadZIP_Batch(directory,zipToRead,rfc_solicitante,testingMode=False):
    separationFolder=''
    if testingMode:
        separationFolder='\\'
    else:
        separationFolder='/'   

    #The first ZIP in the list will be the name and structure of all processes     
    myZip=zipfile.ZipFile(directory+separationFolder+zipToRead,'r')
    #The FIRST zip's file name will be the name of excel file name, like the "Database"
    excel_fileName=excel_fileName=os.path.splitext(os.path.split(myZip.filename)[1])[0]+'.xlsx'
    #Creating the workbook (database)
    #Create the sheets: Ingreso_Egreso,Pago,Resto
    #START - CREATE WORKBOOK FOR ALL POSSIBLE THREADS
    wb=excelpy.Workbook() 
    #let's create the sheets in this order: Emisor, Receptor, Pago_Emisor,Pago_Receptor
    #Sheet1 is the first, so rename it.
    pago_sheet = wb['Sheet']
    pago_sheet.title = 'Emisor'
    wb.create_sheet('Receptor')
    wb.create_sheet('Pago_Emisor')
    wb.create_sheet('Pago_Receptor')
    wb.save(directory+'/'+excel_fileName)
    #dicTableFields is a dictionary with the following structura key:table, value: list of fields
    dicTableFields={}
    #Dictionaries for every kind of "tipo de comprobante"
    #First, get all the columns for all the tables
    for xml in myZip.namelist():
        doc_xml=myZip.open(xml)
        root = ET.parse(doc_xml).getroot()
        for node in root.iter():
            #Column = attribute , Node= Table
            #Get attributes (columns) of current Node (table) 
            #Split the node (table) name because all come as "{something}Node" and {content} is not needed
            #If the number of nodes > 1 then not get its fields, we only want 1 row
            chunk=str(node.tag).split('}')
            tableName=chunk[1] 
            numOfNodes=len(list(node))
            #some "numOfNodes" are Zero, therefore I compare numOfNodes < 2 so I can get 0 and 1
            if (numOfNodes<2) or (numOfNodes>1 and tableName=='Comprobante'):
                chunk=str(node.tag).split('}')
                tableName=chunk[1]  
                #As all the fields will be in one single sheet, it can occur two fields with the same
                #name ex: Rfc and Rfc (Emisor and recipient), then it's needed to add prefix tableName
                for attr in node.attrib:
                    fieldName=tableName+'_'+attr
                    if tableName not in dicTableFields:
                        dicTableFields[tableName]=[fieldName]
                    else:
                        if fieldName not in dicTableFields[tableName]:
                            dicTableFields[tableName].append(fieldName)            
            #End of node iteration 

    #Second, when got all fields from all xml, print them in spread sheet
    lsFields=[] 
    #Start -  Add extra fields in the beginning
    lsFields.append('nombreArchivo')
    lsFields.append('mes')
    #End    - Add extra fields in the beginning

    lsSource=[]
    if len(objControl.lsCustomFields)==0:
        lsSource=dicTableFields   
    else:
        lsSource=objControl.lsCustomFields    

    #I append instead of just assign the list, because I need the column "mes" in the very beginning
    for field in lsSource:
        lsFields.append(field)

    for field in objControl.lsRemove:
        if field in lsFields:
            lsFields.remove(field)     

    #Start - Add extra fields at the end
    lsFields.append('Estatus')
    lsFields.append('Fecha/Hora de Consulta')
    #End - Add extra fields at the end 
    
    #Print all lsFields on excel workbook
    for sheet in wb.sheetnames:
        wb[sheet].append(lsFields)     
    wb.save(directory+'/'+excel_fileName)     

    transformXML_to_XLS(wb,myZip,directory,excel_fileName,lsFields,rfc_solicitante,zipToRead)
   


def transformXML_to_XLS(wb,myZip,directory,excel_fileName,lsFields,rfc_solicitante,nameOfThread):
    #Third, read information and insert where belongs 
    # getroot() : Gets the root of the xml, then use getRoot to get "Comprobante"
    # root.find(.//...): gets any node inside the root, use this to any other node except the root
    contDocs=0
    for xml in myZip.namelist():
        #I need the status for every xml (each cfdi) hence a declare the following fields:
        #Emisor_Rfc,Receptor_Rfc,TimbreFiscalDigital_UUID,Comprobante_Total
        vEmisorRfc=None
        vReceptorRfc=None
        vTimbreFiscal=None
        vComprobanteTotal=None
        #Get field TipoDeComprobante to knwo where sheet to print
        #"Resto" is the default spread sheet
        doc_xml=myZip.open(xml)
        root = ET.parse(doc_xml).getroot()
        lsRfcTable=['Emisor','Receptor']
        sheetPrint='Nada'
        for item in lsRfcTable:
            node=returnFoundNode(root,item)
            if len(node)>0:
                rfc_value=addColumnIfFound(node[0],None,None,'look',['Rfc','rfc'])
                if rfc_value==rfc_solicitante:
                    tipoComprobante=addColumnIfFound(root,None,None,'look',['TipoDeComprobante','tipoDeComprobante'])
                    for possibleValue in ['Ingreso','ingreso','I','i','E','Egreso','egreso','e']:
                        if tipoComprobante==possibleValue:
                            sheetPrint=item
                            break
                    for possibleValue in ['P','p','Pago','pago']:
                        if tipoComprobante==possibleValue:
                            #If "Pago" then it's clear it goes to Pago sheet, so here now divide Pago into
                            #"Pago_Emisor", "Pago_Receptor"
                            if item=='Emisor':
                                sheetPrint='Pago_Emisor'
                            else:
                                sheetPrint='Pago_Receptor'    
                            break
                    if sheetPrint=='Nada': 
                        #If "Nada" it means it goes to Pago, so here now divide Pago into
                        #"Pago_Emisor", "Pago_Receptor"
                        if item=='Emisor':
                            sheetPrint='Pago_Emisor'
                        else:
                            sheetPrint='Pago_Receptor'       

        #Start to read the fields from lsFields=[]
        #Example of a field in lsFields : "Comprobante_Version" -> "tableName_Field"
        #One row per xml
        lsRow=[]
        #The field leads all the insertion
        #Algorith of reading fields
        for field in lsFields:
            #Cases
            if field=='nombreArchivo':
                lsRow.append(xml)
                continue
            if field=='mes': 
                fechaFactura=addColumnIfFound(root,None,None,'look',['Fecha','fecha'])
                monthWord=returnMonthWord(int(fechaFactura.split('-')[1]))
                lsRow.append(monthWord)
                continue
            if field == 'Estatus':
                #Look for this fields and save them to validate status:
                #Emisor_Rfc,Receptor_Rfc,TimbreFiscalDigital_UUID,Comprobante_Total
                #print(f'Obteniendo estado para {vTimbreFiscal}')
                #strStatus=None
                #strStatus=validaEstadoDocumento(vEmisorRfc,vReceptorRfc,vTimbreFiscal,vComprobanteTotal)
                #data=None
                #if strStatus:
                #    data=strStatus['estado']
                #else:
                #    data='No hubo respuesta'
                #Original:
                #lsRow: lsRow.append(data)
                lsRow.append('No status for 802')    
                continue
            if field == 'Fecha/Hora de Consulta': 
                lsRow.append(datetime.datetime.now().strftime(formatDateTime))   
                continue       
            #Rest of cases
            chunks=field.split('_')
            table=chunks[0]
            column=chunks[1]
            if table=='Comprobante':  
                addColumnIfFound(root,column,lsRow,'add',None)
            else:
                #Find the right prefix for table
                lsNode=returnFoundNode(root,table)
                if len(lsNode)==1:
                    addColumnIfFound(lsNode[0],column,lsRow,'add',None)
                elif len(lsNode)>1:
                    #More than 1 table_column found with the same name in XML
                    bTableWithField=False
                    for node in root.findall('.//'+objControl.prefixCFDI+table):
                        if len(node.attrib)>0: 
                            #If this table has attributes, read it, other wise skip it because
                            #if the column doesn't have fields, it means it holds children
                            addColumnIfFound(node,column,lsRow,'add',None)
                            bTableWithField=True
                    if not bTableWithField:
                        lsRow.append(0)    
                        
                else:
                    #No table name found
                    lsRow.append(0)

            #Get values to validate CFDI
            #Look for this fields and save them to validate status:
            #Emisor_Rfc,Receptor_Rfc,TimbreFiscalDigital_UUID,Comprobante_Total
            currentSizeList=None
            currentSizeList=len(lsRow)
            if field=='Emisor_Rfc':
                vEmisorRfc=str(lsRow[currentSizeList-1]) 
                continue  
            if field == 'Receptor_Rfc':
                vReceptorRfc=str(lsRow[currentSizeList-1]) 
                continue
            if field=='TimbreFiscalDigital_UUID':
                vTimbreFiscal=str(lsRow[currentSizeList-1]) 
                continue  
            if field == 'Comprobante_Total':
                vComprobanteTotal=str(lsRow[currentSizeList-1])
                continue           
            #End of field iteration

        #Append the whole xml in a single row            
        wb[sheetPrint].append(lsRow)              
        contDocs+=1
        print(f'{nameOfThread} File done:xml ...{str(contDocs)}') 
        #End of each document (xml) iteration in a zip
        wb.save(directory+'/'+excel_fileName)

    #All xml processed at this point    
    print(f'{nameOfThread} Files processed in ZIP file: {str(contDocs)}') 


          
             
    
    


    
        


   
    


